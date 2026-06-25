"""
ingest.py — universal input layer.

The pipeline only ever consumes ``Candidate`` objects (built by
``Candidate.from_dict``), and ``from_dict`` expects ONE dict shaped like
``config/candidate_schema.json``. Everything in this module exists to turn
*whatever the user uploaded* — JSON / JSONL, CSV / TSV, Excel, or plain text —
into that list of schema-shaped dicts. Nothing downstream (scoring, signals,
integrity, reasoning) changes.

Why this is a separate file: input parsing was previously hard-wired to JSONL in
two places (``schema.stream_candidates`` and the Streamlit app). Keeping the
format logic here means new formats are added in one spot, and the ranking
engine stays format-agnostic.

Supported sources
-----------------
* ``.json``                  — a JSON array OR a single object
* ``.jsonl`` / ``.ndjson``   — one JSON object per line  (also ``.jsonl.gz``)
* ``.csv`` / ``.tsv``        — one candidate per row (flat / dotted / delimited columns)
* ``.xlsx`` / ``.xlsm``      — flat single sheet OR multi-sheet workbook (needs ``openpyxl``)
* ``.txt`` / ``.md`` / other — free text; one candidate per block

The hard part is never the plumbing — it is mapping a *flat* table or *free*
text onto a *deeply nested* schema (profile + variable-length career_history /
education / skills / certifications / languages arrays + a ~23-field
redrob_signals object). The mapping conventions below are what make that
possible; see ``INPUT_FORMATS.md`` for the user-facing version.

Mapping conventions for tabular (CSV / flat Excel) input
--------------------------------------------------------
1. **Bare profile columns** nest automatically under ``profile``:
   ``headline, summary, location, country, years_of_experience, current_title,
   current_company, current_company_size, current_industry, anonymized_name``
   (``name`` is accepted as an alias for ``anonymized_name``).
2. **Bare signal columns** nest automatically under ``redrob_signals``
   (e.g. ``recruiter_response_rate, last_active_date, open_to_work_flag,
   github_activity_score`` …).
3. **Dotted / indexed columns** build arbitrary nesting and arrays — the most
   flexible option:
       ``profile.headline``
       ``redrob_signals.recruiter_response_rate``
       ``redrob_signals.expected_salary_range_inr_lpa.min``
       ``career_history.1.title``   ``career_history.1.company``   (1-based index)
       ``career_history.2.title``   …
4. **Delimited list columns** let a single flat row still carry the arrays.
   Entries are separated by ``;`` and fields within an entry by ``|`` in a
   fixed order:
       ``skills``         →  name | proficiency | endorsements | duration_months
       ``career_history`` →  title | company | industry | duration_months | description
       ``education``      →  institution | degree | field_of_study | start_year | end_year
       ``certifications`` →  name | issuer | year
       ``languages``      →  language | proficiency
   Example skills cell: ``Python|expert|20|36; RAG|advanced|5|8``

Multi-sheet Excel (the lossless option)
---------------------------------------
A workbook with a main sheet (named ``candidates`` / ``profile`` / ``main``, or
simply the first sheet) carrying one flat row per candidate, plus any of these
child sheets keyed by ``candidate_id``:
``career_history, education, skills, certifications, languages, signals``.
Child rows are grouped by ``candidate_id`` and attached as the arrays. The
``signals`` sheet (one row per candidate) is merged into ``redrob_signals``.
"""

from __future__ import annotations

import csv
import gzip
import io
import json
import os
import re
from typing import Any, Iterable

from .schema import Candidate

# --------------------------------------------------------------------------- #
# Field routing tables
# --------------------------------------------------------------------------- #

# Bare column names (no dot) that belong under "profile".
_PROFILE_FIELDS = {
    "anonymized_name", "headline", "summary", "location", "country",
    "years_of_experience", "current_title", "current_company",
    "current_company_size", "current_industry",
}
# Friendly aliases -> canonical profile field.
_PROFILE_ALIASES = {
    "name": "anonymized_name",
    "full_name": "anonymized_name",
    "title": "current_title",
    "company": "current_company",
    "industry": "current_industry",
    "company_size": "current_company_size",
    "years": "years_of_experience",
    "yoe": "years_of_experience",
    "experience_years": "years_of_experience",
}
# Bare column names (no dot) that belong under "redrob_signals".
_SIGNAL_FIELDS = {
    "profile_completeness_score", "signup_date", "last_active_date",
    "open_to_work_flag", "profile_views_received_30d",
    "applications_submitted_30d", "recruiter_response_rate",
    "avg_response_time_hours", "connection_count", "endorsements_received",
    "notice_period_days", "preferred_work_mode", "willing_to_relocate",
    "github_activity_score", "search_appearance_30d", "saved_by_recruiters_30d",
    "interview_completion_rate", "offer_acceptance_rate", "verified_email",
    "verified_phone", "linkedin_connected",
}
# Names whose values should be coerced to bool when they look truthy/falsy.
_BOOL_FIELDS = {
    "open_to_work_flag", "willing_to_relocate", "verified_email",
    "verified_phone", "linkedin_connected", "is_current",
}

# Field order for the pipe-delimited list columns (single-flat-sheet mode).
_LIST_SPECS: dict[str, list[str]] = {
    "skills": ["name", "proficiency", "endorsements", "duration_months"],
    "career_history": ["title", "company", "industry", "duration_months", "description"],
    "education": ["institution", "degree", "field_of_study", "start_year", "end_year"],
    "certifications": ["name", "issuer", "year"],
    "languages": ["language", "proficiency"],
}
# Child-sheet names recognised in multi-sheet workbooks.
_CHILD_SHEETS = {"career_history", "education", "skills", "certifications", "languages"}
_MAIN_SHEET_NAMES = {"candidates", "candidate", "profile", "profiles", "main", "sheet1"}

_INT_RE = re.compile(r"^[+-]?\d+$")
_FLOAT_RE = re.compile(r"^[+-]?(\d+\.\d*|\.\d+|\d+)([eE][+-]?\d+)?$")


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #

def load_any(source: Any, *, filename: str | None = None,
             fmt: str | None = None) -> list[Candidate]:
    """Load candidates from any supported source.

    ``source`` may be a filesystem path (str) or raw bytes/str (e.g. from a
    Streamlit upload or a paste box). When passing bytes/str, also pass
    ``filename`` so the format can be detected from the extension.

    ``fmt`` optionally overrides auto-detection. Accepted values:
    ``"json"``, ``"jsonl"``, ``"csv"``, ``"tsv"``, ``"xlsx"``, ``"text"``
    (``None`` or ``"auto"`` = detect from extension, or sniff if no filename).
    """
    return [
        Candidate.from_dict(d)
        for d in records_from_any(source, filename=filename, fmt=fmt)
    ]


def records_from_any(source: Any, *, filename: str | None = None,
                     fmt: str | None = None) -> list[dict[str, Any]]:
    """Return schema-shaped dicts from any supported source (no Candidate wrap).

    ``source`` may be a filesystem path, raw bytes, raw text (e.g. a paste box),
    or a file-like object. A ``str`` is treated as a path only if it exists on
    disk; otherwise it is treated as raw text content.
    """
    # Resolve raw bytes + an optional name to detect the format from.
    path_name = ""
    if isinstance(source, str) and os.path.exists(source):  # path on disk
        raw_bytes = _read_path_bytes(source)
        path_name = source
    elif isinstance(source, str):                           # raw text content
        raw_bytes = source.encode("utf-8")
    elif isinstance(source, (bytes, bytearray)):
        raw_bytes = bytes(source)
    elif hasattr(source, "read"):                           # file-like object
        data = source.read()
        raw_bytes = data if isinstance(data, (bytes, bytearray)) else str(data).encode("utf-8")
        path_name = getattr(source, "name", "") or ""
    else:
        raw_bytes = str(source).encode("utf-8")

    name = (filename or path_name or "").lower()

    # Resolve the effective format: explicit override > extension > content sniff.
    chosen = (fmt or "").lower().strip()
    if chosen in ("", "auto"):
        chosen = _detect_format(name, raw_bytes)

    return _records_for_format(chosen, raw_bytes)


def _detect_format(name: str, raw_bytes: bytes) -> str:
    """Pick a format code from the filename extension, or sniff the content."""
    ext = _ext(name)
    if name.endswith(".jsonl.gz") or ext == ".gz":
        return "jsonl_gz"
    if ext in (".jsonl", ".ndjson"):
        return "jsonl"
    if ext == ".json":
        return "json"
    if ext == ".csv":
        return "csv"
    if ext == ".tsv":
        return "tsv"
    if ext in (".xlsx", ".xlsm", ".xls"):
        return "xlsx"
    if ext in (".txt", ".md"):
        return "text"
    # No / unknown extension (e.g. pasted text): sniff from the content.
    return _sniff_format(_text(raw_bytes))


def _records_for_format(fmt: str, raw_bytes: bytes) -> list[dict[str, Any]]:
    """Dispatch raw bytes to the right parser for a resolved format code."""
    if fmt == "jsonl_gz":
        text = gzip.decompress(raw_bytes).decode("utf-8", errors="replace")
        return _records_from_json_text(text)
    if fmt == "jsonl":
        return _records_from_json_text(_text(raw_bytes), force_lines=True)
    if fmt == "json":
        # Handles a JSON array, a single object, OR (fallback) one-object-per-line.
        return _records_from_json_text(_text(raw_bytes))
    if fmt == "csv":
        return _records_from_csv(_text(raw_bytes), delimiter=",")
    if fmt == "tsv":
        return _records_from_csv(_text(raw_bytes), delimiter="\t")
    if fmt == "xlsx":
        return _records_from_excel(raw_bytes)
    # "text" and any unknown code → treat as free text.
    return _records_from_text(_text(raw_bytes))


def _sniff_format(text: str) -> str:
    """Best-effort format guess for content with no filename (e.g. pasted text)."""
    stripped = text.lstrip()
    if not stripped:
        return "text"
    lines = [ln for ln in text.splitlines() if ln.strip()][:5]
    # Multiple standalone JSON objects (one per line) → JSONL.
    if len(lines) >= 2 and all(_looks_like_json_object(ln) for ln in lines):
        return "jsonl"
    # A JSON array, or a single (possibly multi-line) object.
    if stripped[0] in "[{":
        return "json"
    # Header-ish first line with delimiters → CSV/TSV.
    first = lines[0] if lines else ""
    if "\t" in first:
        return "tsv"
    if "," in first and not first.strip().endswith(":"):
        return "csv"
    return "text"


def _looks_like_json_object(line: str) -> bool:
    s = line.strip().rstrip(",")
    if not (s.startswith("{") and s.endswith("}")):
        return False
    try:
        return isinstance(json.loads(s), dict)
    except json.JSONDecodeError:
        return False


# --------------------------------------------------------------------------- #
# JSON / JSONL
# --------------------------------------------------------------------------- #

def _records_from_json_text(text: str, *, force_lines: bool = False) -> list[dict[str, Any]]:
    stripped = text.lstrip()
    # A JSON array or single object (only if not explicitly a .jsonl file).
    if not force_lines and stripped[:1] in ("[", "{"):
        try:
            obj = json.loads(text)
            if isinstance(obj, list):
                return [o for o in obj if isinstance(o, dict)]
            if isinstance(obj, dict):
                # could be {"candidates": [...]} or a single record
                for key in ("candidates", "records", "data", "items"):
                    if isinstance(obj.get(key), list):
                        return [o for o in obj[key] if isinstance(o, dict)]
                return [obj]
        except json.JSONDecodeError:
            pass  # fall back to line-by-line
    out: list[dict[str, Any]] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            d = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(d, dict):
            out.append(d)
    return out


# --------------------------------------------------------------------------- #
# CSV / TSV  (flat one-row-per-candidate)
# --------------------------------------------------------------------------- #

def _records_from_csv(text: str, *, delimiter: str = ",") -> list[dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = [r for r in reader]
    records = [_row_to_record(r) for r in rows]
    _autofill_ids(records)
    return records


# --------------------------------------------------------------------------- #
# Excel  (flat single sheet OR multi-sheet workbook)
# --------------------------------------------------------------------------- #

def _records_from_excel(raw_bytes: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - depends on environment
        raise ImportError(
            "Reading .xlsx requires openpyxl. Install it with:\n"
            "    pip install openpyxl\n"
            "(CSV, TSV, JSON and plain-text inputs need no extra packages.)"
        ) from e

    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    sheets = {ws.title.strip().lower(): ws for ws in wb.worksheets}

    # Decide layout: multi-sheet if any recognised child/signals sheet exists.
    has_children = bool(_CHILD_SHEETS & set(sheets)) or "signals" in sheets
    main_key = _pick_main_sheet(sheets, has_children)
    main_rows = _sheet_to_dicts(sheets[main_key]) if main_key else []

    if not has_children:
        # Single flat sheet → same path as CSV.
        records = [_row_to_record(r) for r in main_rows]
        _autofill_ids(records)
        return records

    # Multi-sheet: build base records from the main sheet, then attach children.
    records = [_row_to_record(r) for r in main_rows]
    _autofill_ids(records)
    by_id = {r.get("candidate_id"): r for r in records}

    for child in _CHILD_SHEETS:
        if child not in sheets:
            continue
        for row in _sheet_to_dicts(sheets[child]):
            cid = _first_present(row, ("candidate_id", "candidate id", "id"))
            rec = by_id.get(str(cid)) if cid is not None else None
            if rec is None:
                continue
            entry = {
                k: _coerce(k, v)
                for k, v in row.items()
                if k and k.lower() not in ("candidate_id", "candidate id", "id") and v not in (None, "")
            }
            if entry:
                rec.setdefault(child, []).append(entry)

    if "signals" in sheets:
        for row in _sheet_to_dicts(sheets["signals"]):
            cid = _first_present(row, ("candidate_id", "candidate id", "id"))
            rec = by_id.get(str(cid)) if cid is not None else None
            if rec is None:
                continue
            sig = rec.setdefault("redrob_signals", {})
            for k, v in row.items():
                if not k or k.lower() in ("candidate_id", "candidate id", "id") or v in (None, ""):
                    continue
                sig[k.strip()] = _coerce(k, v)

    return records


def _pick_main_sheet(sheets: dict[str, Any], has_children: bool) -> str | None:
    for name in _MAIN_SHEET_NAMES:
        if name in sheets:
            return name
    # Otherwise the first sheet that is not a recognised child / signals sheet.
    for name in sheets:
        if name not in _CHILD_SHEETS and name != "signals":
            return name
    # Degenerate workbook that is *only* child sheets: use the first sheet.
    return next(iter(sheets), None)


def _sheet_to_dicts(ws) -> list[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    cols = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
    out: list[dict[str, Any]] = []
    for r in rows:
        if r is None or all(v is None or v == "" for v in r):
            continue
        out.append({cols[i]: r[i] for i in range(min(len(cols), len(r)))})
    return out


# --------------------------------------------------------------------------- #
# Free text (.txt / .md / unknown)
# --------------------------------------------------------------------------- #

# A line that is only dashes / asterisks / equals separates candidate blocks.
_BLOCK_SEP = re.compile(r"^\s*([-*=_]\s*){3,}\s*$")
# Leading "key: value" lines we recognise inside a block.
_TXT_KEY_MAP = {
    "candidate_id": ("candidate_id", None),
    "id": ("candidate_id", None),
    "name": ("anonymized_name", "profile"),
    "headline": ("headline", "profile"),
    "title": ("current_title", "profile"),
    "current_title": ("current_title", "profile"),
    "company": ("current_company", "profile"),
    "current_company": ("current_company", "profile"),
    "location": ("location", "profile"),
    "country": ("country", "profile"),
    "industry": ("current_industry", "profile"),
    "years": ("years_of_experience", "profile"),
    "years_of_experience": ("years_of_experience", "profile"),
}


def _records_from_text(text: str) -> list[dict[str, Any]]:
    blocks = _split_text_blocks(text)
    records: list[dict[str, Any]] = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        records.append(_text_block_to_record(block))
    _autofill_ids(records)
    return records


def _split_text_blocks(text: str) -> list[str]:
    lines = text.splitlines()
    # Explicit separators take priority.
    if any(_BLOCK_SEP.match(ln) for ln in lines):
        blocks, cur = [], []
        for ln in lines:
            if _BLOCK_SEP.match(ln):
                blocks.append("\n".join(cur))
                cur = []
            else:
                cur.append(ln)
        blocks.append("\n".join(cur))
        return [b for b in blocks if b.strip()]
    # Else split on blank-line gaps of 2+ lines (paragraph = candidate).
    if re.search(r"\n\s*\n\s*\n", text):
        return [b for b in re.split(r"\n\s*\n\s*\n+", text) if b.strip()]
    # No delimiters at all → the whole file is a single candidate.
    return [text]


def _text_block_to_record(block: str) -> dict[str, Any]:
    rec: dict[str, Any] = {"profile": {}}
    for line in block.splitlines():
        m = re.match(r"^\s*([A-Za-z_ ]{2,30}?)\s*:\s*(.+?)\s*$", line)
        if not m:
            continue
        key = m.group(1).strip().lower().replace(" ", "_")
        if key in _TXT_KEY_MAP:
            field, parent = _TXT_KEY_MAP[key]
            val = _coerce(field, m.group(2))
            if parent == "profile":
                rec["profile"][field] = val
            else:
                rec[field] = val
    # The whole block is also stored as the summary so the semantic encoder
    # (which reads headline + summary + career text) has the full text to work
    # with — this is what keeps a bare résumé rankable.
    rec["profile"].setdefault("summary", block)
    return rec


# --------------------------------------------------------------------------- #
# Flat row -> nested record  (shared by CSV and single-sheet Excel)
# --------------------------------------------------------------------------- #

def _row_to_record(row: dict[str, Any]) -> dict[str, Any]:
    rec: dict[str, Any] = {}
    for raw_key, raw_val in row.items():
        if raw_key is None:
            continue
        key = str(raw_key).strip()
        if key == "" or raw_val is None or (isinstance(raw_val, str) and raw_val.strip() == ""):
            continue
        low = key.lower()

        # 1) Dotted / indexed path → build nested structure.
        if "." in key:
            _assign_path(rec, key, raw_val)
            continue

        # 2) candidate_id stays top-level.
        if low in ("candidate_id", "candidate id", "id"):
            rec["candidate_id"] = str(raw_val).strip()
            continue

        # 3) Delimited list columns (skills, career_history, …).
        if low in _LIST_SPECS:
            rec[low] = _parse_list_column(low, str(raw_val))
            continue

        # 4) Bare profile field (with alias support).
        canon = _PROFILE_ALIASES.get(low, low)
        if canon in _PROFILE_FIELDS:
            rec.setdefault("profile", {})[canon] = _coerce(canon, raw_val)
            continue

        # 5) Bare signal field.
        if low in _SIGNAL_FIELDS:
            rec.setdefault("redrob_signals", {})[low] = _coerce(low, raw_val)
            continue

        # 6) Unknown bare column: ignore (keeps the schema clean). Use a dotted
        #    header like ``profile.my_field`` if you need to force it in.
    return rec


def _assign_path(root: dict[str, Any], dotted: str, value: Any) -> None:
    """Assign ``value`` into ``root`` following a dotted/indexed path.

    Integer segments (1-based) index into lists; string segments index dicts.
    e.g. ``career_history.2.title`` -> root['career_history'][1]['title'].
    """
    parts = [p.strip() for p in dotted.split(".") if p.strip() != ""]
    if not parts:
        return
    cur: Any = root
    for i, part in enumerate(parts):
        last = i == len(parts) - 1
        is_index = _INT_RE.match(part) is not None
        nxt_is_index = (not last) and _INT_RE.match(parts[i + 1]) is not None

        if is_index:
            idx = int(part) - 1
            if idx < 0:
                idx = 0
            if not isinstance(cur, list):
                # parent created the wrong container; coerce silently
                return
            while len(cur) <= idx:
                cur.append({} if not nxt_is_index else [])
            if last:
                # rare: a list of scalars
                cur[idx] = _coerce(parts[i - 1] if i else part, value)
            else:
                if cur[idx] is None:
                    cur[idx] = [] if nxt_is_index else {}
                cur = cur[idx]
        else:
            if not isinstance(cur, dict):
                return
            if last:
                cur[part] = _coerce(part, value)
            else:
                want_list = nxt_is_index
                existing = cur.get(part)
                if existing is None or (want_list and not isinstance(existing, list)) \
                        or (not want_list and not isinstance(existing, dict)):
                    cur[part] = [] if want_list else {}
                cur = cur[part]


def _parse_list_column(name: str, cell: str) -> list[dict[str, Any]]:
    """Parse a ``;``-separated, ``|``-fielded list cell into list-of-dicts."""
    spec = _LIST_SPECS[name]
    out: list[dict[str, Any]] = []
    for chunk in cell.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        fields = [f.strip() for f in chunk.split("|")]
        entry: dict[str, Any] = {}
        for i, field_name in enumerate(spec):
            if i < len(fields) and fields[i] != "":
                entry[field_name] = _coerce(field_name, fields[i])
        if entry:
            out.append(entry)
    return out


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def _coerce(field_name: str, value: Any) -> Any:
    """Best-effort type coercion. Dates are kept as strings (schema wants that)."""
    if not isinstance(value, str):
        return value  # openpyxl already gives native int/float/bool/datetime
    s = value.strip()
    if s == "":
        return s
    low_field = field_name.lower()
    if low_field in _BOOL_FIELDS:
        if s.lower() in ("true", "yes", "y", "1"):
            return True
        if s.lower() in ("false", "no", "n", "0"):
            return False
    # Never numeric-coerce identifier-ish or date-ish strings.
    if "_id" in low_field or low_field.endswith("date") or low_field in ("start_date", "end_date"):
        return s
    if _INT_RE.match(s):
        try:
            return int(s)
        except ValueError:
            return s
    if _FLOAT_RE.match(s):
        try:
            return float(s)
        except ValueError:
            return s
    return s


def _autofill_ids(records: list[dict[str, Any]]) -> None:
    """Give every record a stable candidate_id if it lacks one."""
    for i, rec in enumerate(records):
        cid = rec.get("candidate_id")
        if not cid or not str(cid).strip():
            rec["candidate_id"] = f"CAND_AUTO_{i:07d}"


def _first_present(row: dict[str, Any], keys: Iterable[str]) -> Any:
    low = {str(k).strip().lower(): v for k, v in row.items()}
    for k in keys:
        if k in low and low[k] not in (None, ""):
            return low[k]
    return None


def _read_path_bytes(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()


def _text(raw_bytes: bytes) -> str:
    return raw_bytes.decode("utf-8", errors="replace")


def _ext(name: str) -> str:
    return os.path.splitext(name)[1].lower()
